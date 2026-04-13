"""Excel workbook builder for apartment listings."""

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .config import EXCEL_FILE, SITE_SHEET_NAMES, TODAY

# ──────────────────────────────────────────────────────────────────────────────
# COLUMN DEFINITIONS
# ──────────────────────────────────────────────────────────────────────────────

COLUMNS = [
    "Source", "Neighborhood", "Price", "Beds", "Baths",
    "Sqft", "Address", "URL", "Notes", "Date Found",
]
COL_WIDTHS = [14, 20, 12, 8, 8, 10, 35, 50, 30, 14]

LISTING_KEYS = [
    "source", "neighborhood", "price", "beds", "baths",
    "sqft", "address", "url", "notes", "date_found",
]

URL_COL = COLUMNS.index("URL") + 1  # 1-indexed

# ──────────────────────────────────────────────────────────────────────────────
# STYLES
# ──────────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

BODY_FONT = Font(name="Arial", size=10)
BODY_FONT_BOLD = Font(name="Arial", size=10, bold=True)
ALT_ROW_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
NEW_ROW_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _style_header(ws):
    """Apply header styling, column widths, freeze panes, and auto-filter."""
    for col_idx, (name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def _append_listings(ws, listings, is_new=True):
    """Append listing rows to a worksheet."""
    for listing in listings:
        row_num = ws.max_row + 1
        values = [listing[k] for k in LISTING_KEYS]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = BODY_FONT

            if col_idx == URL_COL and val:
                cell.hyperlink = val
                cell.font = LINK_FONT

            if is_new:
                cell.fill = NEW_ROW_FILL
            elif row_num % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Expand auto-filter range
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"


def _reapply_alternating_fill(ws):
    """Re-apply row fills: green for today's new listings, alternating blue otherwise."""
    for row_num in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row_num, column=len(COLUMNS))
        is_today = str(date_cell.value) == TODAY

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_num, column=col_idx)

            if is_today:
                cell.fill = NEW_ROW_FILL
            elif row_num % 2 == 0:
                cell.fill = ALT_ROW_FILL
            else:
                cell.fill = NO_FILL

            if col_idx == URL_COL and cell.value:
                cell.font = LINK_FONT


# ──────────────────────────────────────────────────────────────────────────────
# PUBLIC API
# ──────────────────────────────────────────────────────────────────────────────

def build_excel(all_new_listings: list[dict]) -> None:
    """Build or update the Excel workbook with new listings."""
    # Load existing or create new
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    by_source = {}
    for listing in all_new_listings:
        by_source.setdefault(listing["source"], []).append(listing)

    # ── Per-site sheets ──
    for site_name, sheet_name in SITE_SHEET_NAMES.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            _style_header(ws)
        else:
            ws = wb[sheet_name]

        site_listings = by_source.get(site_name, [])
        if site_listings:
            _append_listings(ws, site_listings, is_new=True)
        _reapply_alternating_fill(ws)

    # ── All Listings sheet ──
    all_sheet_name = "📋 All Listings"
    if all_sheet_name not in wb.sheetnames:
        ws_all = wb.create_sheet(all_sheet_name)
        _style_header(ws_all)
    else:
        ws_all = wb[all_sheet_name]

    if all_new_listings:
        _append_listings(ws_all, all_new_listings, is_new=True)
    _reapply_alternating_fill(ws_all)

    # ── Summary sheet (recreated every run) ──
    summary_name = "📊 Summary"
    if summary_name in wb.sheetnames:
        del wb[summary_name]
    ws_sum = wb.create_sheet(summary_name, 0)

    sum_headers = ["Site", "New Today", "Total Listings"]
    sum_widths = [20, 14, 16]
    for col_idx, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        cell = ws_sum.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = w
    ws_sum.freeze_panes = "A2"
    ws_sum.auto_filter.ref = "A1:C1"

    for row_idx, (site_name, sheet_name) in enumerate(SITE_SHEET_NAMES.items(), 2):
        new_count = len(by_source.get(site_name, []))
        total_count = max(wb[sheet_name].max_row - 1, 0) if sheet_name in wb.sheetnames else 0

        ws_sum.cell(row=row_idx, column=1, value=site_name).font = BODY_FONT
        ws_sum.cell(row=row_idx, column=2, value=new_count).font = BODY_FONT
        ws_sum.cell(row=row_idx, column=3, value=total_count).font = BODY_FONT

        if row_idx % 2 == 0:
            for c in range(1, 4):
                ws_sum.cell(row=row_idx, column=c).fill = ALT_ROW_FILL

    # Totals row
    total_row = len(SITE_SHEET_NAMES) + 2
    ws_sum.cell(row=total_row, column=1, value="TOTAL").font = BODY_FONT_BOLD
    ws_sum.cell(row=total_row, column=2, value=len(all_new_listings)).font = BODY_FONT_BOLD
    total_all = max(wb[all_sheet_name].max_row - 1, 0) if all_sheet_name in wb.sheetnames else 0
    ws_sum.cell(row=total_row, column=3, value=total_all).font = BODY_FONT_BOLD

    # ── Reorder: summary first ──
    wb.move_sheet(summary_name, offset=-(wb.sheetnames.index(summary_name)))

    wb.save(EXCEL_FILE)
    print(f"\n  ✓ Excel saved: {EXCEL_FILE}")
